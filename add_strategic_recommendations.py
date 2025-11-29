#!/usr/bin/env python3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load the workbook
wb = load_workbook('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
ws = wb.active

def get_strategic_recommendation(vendor_name):
    """
    Classify vendor using STRICT rules:
    - TERMINATE: Travel, hotels, restaurants, catering, events, parking, non-critical
    - CONSOLIDATE: Multiple SaaS tools, recruiting agencies, IT consultancies, cloud platforms
    - OPTIMIZE: Mission-critical SaaS (AWS, Salesforce, Microsoft, Adobe, etc.)
    """
    vendor_lower = vendor_name.lower()

    # TERMINATE - Travel, hotels, restaurants, catering, events, local vendors, non-critical
    terminate_keywords = [
        # Travel & hospitality
        'hotel', 'resort', 'accommodation', 'inn', 'pastoria', 'intercontinental',
        'radisson', 'hilton', 'trocadero', 'zonar', 'laguna', 'winery',
        # Restaurants & food
        'restaurant', 'cafe', 'coffee', 'catering', 'kitchen', 'dining', 'food',
        'bar', 'tattu', 'gaucho', 'mesa verde', 'pret a manger', 'bakery',
        'cupcake', 'saloon', 'italian', 'del posto', 'harissa', 'pepe',
        # Events & entertainment
        'event', 'comedy', 'entertainment', 'escape art', 'paint&wine', 'paint & fun',
        'djs for u', 'blink events', 'rishi events', 'urbani eventi',
        # Parking & transport
        'parking', 'garage', 'golubica', 'firule', 'uber', 'wolt',
        # Travel services
        'travel', 'tour', 'airline', 'croatia airlines', 'hahn air',
        # Local/one-off vendors
        'student packers', 'office move', 'moving', 'relocation',
        # Personal/non-essential
        'gym', 'fitness', 'sports club', 'recreation', 'cycle gap', 'athlete service',
        'wine', 'istra wine', 'vivat fina',
        # Retail/shopping (non-essential)
        'pink ribbon', 'regency hampers', 'plant man', 'notino', 'freepik',
        'snappy snaps', 'vistaprint', 'gift', 'hampers', 'flower', 'floom',
        # Individual contractors
        'john smith', 'susan lee', 'george anchor', 'fabiola', 'stipe piric', 'ansar madovic',
        # Unclear/non-critical
        'smell', 'decoration', 'canteen', 'vending',
    ]

    for keyword in terminate_keywords:
        if keyword in vendor_lower:
            return 'Terminate'

    # OPTIMIZE - Mission-critical SaaS platforms and infrastructure
    optimize_keywords = [
        # Cloud infrastructure (mission-critical)
        'aws', 'amazon web services', 'microsoft', 'azure', 'google cloud',
        # Major SaaS platforms
        'salesforce', 'adobe', 'atlassian', 'figma', 'slack',
        'docusign', 'smartsheet', 'workato', 'zapier',
        # Development tools
        'jetbrains', 'npm', 'github', 'gitlab',
        # Critical business platforms
        'hubspot', 'linkedin', 'ariba', 'kimble', 'planful',
        # Infrastructure & monitoring
        'solarwinds', 'uptime robot', 'papertrail', 'lastpass',
        # Training platforms
        'pluralsight', 'interaction design foundation',
        # Big 4 / Major professional services
        'bdo llp', 'grant thornton', 'pricewaterhouse', 'deloitte', 'kpmg', 'ey',
        'houlihan lokey', 'crowe horwath',
        # Essential IT services
        'infosys', 'dhl', 'fedex',
        # Primary travel management
        'navan', 'tripactions',
        # Primary real estate
        'cbre', 'jones lang lasalle',
        # Core HR/benefits
        'mercer limited', 'benefit systems', 'pluxee', 'sodexo',
        'granttree limited',  # R&D tax credits
    ]

    for keyword in optimize_keywords:
        if keyword in vendor_lower:
            return 'Optimize'

    # CONSOLIDATE - Everything else (multiple tools, agencies, overlapping services)
    # This includes:
    # - Multiple SaaS tools in same category
    # - Multiple recruiting agencies
    # - Multiple IT consultancies
    # - Multiple cloud platforms
    # - Multiple analytics/HR tools
    # - Multiple legal firms
    # - Multiple accounting firms
    # - Multiple insurance providers
    # - Multiple telecom providers
    # - Multiple office space providers
    # - All other business services

    return 'Consolidate'

# Check if "Strategic Recommendation" column already exists
header_row = list(ws[1])
column_headers = [cell.value for cell in header_row]

# Find or create the Strategic Recommendation column
if 'Strategic Recommendation' in column_headers:
    rec_col_idx = column_headers.index('Strategic Recommendation') + 1
    print("Found existing 'Strategic Recommendation' column, updating it...")
else:
    # Add new column after the existing columns
    rec_col_idx = len(column_headers) + 1
    rec_col_letter = ws.cell(row=1, column=rec_col_idx).column_letter

    # Add header
    header_cell = ws.cell(row=1, column=rec_col_idx)
    header_cell.value = 'Strategic Recommendation'
    header_cell.font = Font(bold=True, size=12, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_cell.alignment = Alignment(horizontal='left', vertical='center')

    print(f"Created new 'Strategic Recommendation' column at position {rec_col_letter}")

# Process each vendor and add recommendation
recommendations_count = {'Terminate': 0, 'Consolidate': 0, 'Optimize': 0}

for row_num in range(2, ws.max_row + 1):
    vendor_cell = ws.cell(row=row_num, column=1)  # Column A has vendor names
    vendor_name = vendor_cell.value

    if vendor_name:
        recommendation = get_strategic_recommendation(vendor_name)

        # Write recommendation to the new column
        rec_cell = ws.cell(row=row_num, column=rec_col_idx)
        rec_cell.value = recommendation
        rec_cell.alignment = Alignment(horizontal='left', vertical='top')

        # Color code the recommendations
        if recommendation == 'Terminate':
            rec_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
            rec_cell.font = Font(color="CC0000", bold=True)  # Dark red text
        elif recommendation == 'Consolidate':
            rec_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")  # Light orange
            rec_cell.font = Font(color="CC6600", bold=True)  # Dark orange text
        elif recommendation == 'Optimize':
            rec_cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")  # Light green
            rec_cell.font = Font(color="0D652D", bold=True)  # Dark green text

        recommendations_count[recommendation] += 1

# Adjust column width
ws.column_dimensions[ws.cell(row=1, column=rec_col_idx).column_letter].width = 25

# Save the updated workbook
output_file = '/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx'
wb.save(output_file)

print(f"\n✓ Updated spreadsheet saved: {output_file}")
print(f"\nRecommendations Summary:")
print(f"  Terminate:    {recommendations_count['Terminate']} vendors")
print(f"  Consolidate:  {recommendations_count['Consolidate']} vendors")
print(f"  Optimize:     {recommendations_count['Optimize']} vendors")
print(f"  Total:        {sum(recommendations_count.values())} vendors")
