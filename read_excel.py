#!/usr/bin/env python3
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
ws = wb.active

# Print all rows
for row in ws.iter_rows(values_only=True):
    print(row)
