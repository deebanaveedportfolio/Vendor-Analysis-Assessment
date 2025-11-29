#!/usr/bin/env python3
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
ws = wb.active

def get_vendor_recommendation(vendor_name, vendor_lower):
    """Generate recommendation: Terminate, Consolidate, or Optimize"""

    # TERMINATE - Non-essential or easily replaceable services
    terminate_keywords = [
        'individual contractor', 'john smith', 'susan lee', 'george anchor', 'fabiola thistlewhaite',
        'stipe piric', 'ansar madovic',  # Individual contractors
        'pink ribbon shop', 'regency hampers', 'cupcake central', 'the plant man',  # Non-essential retail
        'djs for u', 'paint & fun', 'paint&wine', 'lajnap comedy',  # Entertainment/events (non-core)
        'escape art', 'magic mountain saloon', 'pepe\'s italian',  # Recreation venues
        'gym4you', 'athlete service', 'friends sports club', 'p s recreation',  # Gym memberships
        'chamiers recreation', 'the cycle gap',  # Recreation
        'istra wine', 'vivat fina vina',  # Wine/beverages (non-essential)
        'notino s.r.o.', 'freepik company',  # Can use free/cheaper alternatives
    ]

    # CONSOLIDATE - Multiple vendors doing similar things
    consolidate_keywords = [
        # Office space - consolidate multiple providers
        'office space', 'coworking', 'wework', 'tog uk', 'common desk', 'innovent spaces',
        'work easy space', 'big frontier', 'gpt space', 'platinum office',

        # Cloud services - consolidate to primary provider
        'cloud', 'cloudcrossing', 'cloud technology solutions',

        # Recruitment - consolidate agencies
        'recruitment', 'cedar recruitment', 'mason frank', 'technet it recruitment',
        'integrated personnel',

        # Hotels - use corporate travel platform
        'hotel', 'resort', 'inter continental', 'radisson', 'puducherry backwater',
        'trocadero', 'hilton garden', 'president hotel', 'marvie hotel', 'obiteljski hoteli',

        # Catering - consolidate food services
        'catering', 'food services', 'city pantry', 'my foodiverse', 'lunch nutrition',
        'kat\'s kitchen', 'soho kitchen', 'the cook kitchen', 'taste of health',

        # Legal firms - consolidate to primary counsel
        'legal', 'solicitor', 'law firm', 'bisley law', 'quadrant law', 'curzon green',
        'thomas mansfield', 'landu law', 'induslaw',

        # Accounting firms - consolidate to primary firm
        'chartered accountants', 'collards', 'mcburneys', 'n s shastri',

        # Insurance - consolidate policies
        'insurance', 'bupa australia', 'cigna sg', 'cici prudential',

        # Telecoms - consolidate providers
        'telecommunications', 'telekom', 'vodafone', 't-mobile', 'starhub', 'telemach',

        # Parking - consolidate parking services
        'parking', 'golubica parking', 'garaå¾a firule',

        # Event planning - consolidate event services
        'event', 'blink events', 'event ors', 'urbani eventi', 'rishi events',

        # Corporate services - consolidate to primary provider
        'acclime corporate', 'acclime usa', 'intertrust singapore',

        # Student accommodation - consolidate
        'studentski centar', 'student accommodation',
    ]

    # OPTIMIZE - Critical/core services to keep but optimize costs
    optimize_keywords = [
        # Core platforms (keep but negotiate)
        'salesforce', 'aws', 'amazon web services', 'microsoft', 'google ireland',
        'hubspot', 'linkedin', 'workato', 'atlassian', 'figma', 'adobe',
        'docusign', 'smartsheet', 'trello', 'slack', 'zapier',

        # Critical professional services
        'bdo llp', 'grant thornton', 'pricewaterhousecoopers', 'rsm uk corporate',
        'houlihan lokey', 'crowe horwath',  # Major accounting/advisory firms

        # Core IT/Engineering
        'infosys', 'jetbrains', 'ariba', 'kimble', 'pluralsight',

        # Essential infrastructure
        'dhl', 'fedex', 'british telecommunications',

        # Primary travel management
        'navan (tripactions inc)', 'navan, inc',

        # Primary real estate
        'cbre limited', 'jones lang lasalle',

        # Employee benefits platforms
        'benefit systems', 'mercer limited', 'pluxee india',

        # Core business services
        'sodexo', 'granttree limited',  # R&D tax credits

        # Critical software tools
        'lastpass', 'solarwinds', 'uptime robot', 'papertrail',
    ]

    # Check terminate conditions
    for keyword in terminate_keywords:
        if keyword in vendor_lower:
            return 'Terminate'

    # Check consolidate conditions
    for keyword in consolidate_keywords:
        if keyword in vendor_lower:
            return 'Consolidate'

    # Check optimize conditions
    for keyword in optimize_keywords:
        if keyword in vendor_lower:
            return 'Optimize'

    # Default categorization based on vendor type
    if any(word in vendor_lower for word in ['d.o.o.', 'j.d.o.o.', 'obrt']):
        # Many Croatian local businesses - likely consolidate or terminate
        if any(word in vendor_lower for word in ['restaurant', 'bar', 'cafe', 'coffee', 'bakery']):
            return 'Consolidate'  # Food/beverage vendors
        elif any(word in vendor_lower for word in ['grocery', 'retail', 'shop', 'store']):
            return 'Consolidate'  # Retail
        else:
            return 'Optimize'  # Other local services

    # Generic business services providers
    if 'business services provider' in get_description(vendor_name):
        return 'Consolidate'

    # Default to Optimize for undefined vendors
    return 'Optimize'

def get_description(vendor_name):
    """Get vendor description for context"""
    vendor_lower = vendor_name.lower()
    if 'hotel' in vendor_lower or 'resort' in vendor_lower:
        return 'hotel services'
    elif 'catering' in vendor_lower or 'food' in vendor_lower or 'kitchen' in vendor_lower:
        return 'catering services'
    elif 'law' in vendor_lower or 'legal' in vendor_lower or 'solicitor' in vendor_lower:
        return 'legal services'
    else:
        return 'business services provider'

# Print table
print("| Vendor Name | Recommendation |")
print("|-------------|----------------|")

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        vendor_name = row[0]
        vendor_lower = vendor_name.lower()
        recommendation = get_vendor_recommendation(vendor_name, vendor_lower)
        print(f"| {vendor_name} | {recommendation} |")
