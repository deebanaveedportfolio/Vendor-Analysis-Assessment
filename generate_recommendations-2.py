#!/usr/bin/env python3
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
ws = wb.active

# Vendor descriptions dictionary (imported from generate_descriptions.py)
VENDOR_DESCRIPTIONS = {
    'salesforce uk ltd-uk': 'Cloud-based CRM and sales automation platform',
    'navan (tripactions inc)': 'Corporate travel and expense management platform',
    'bdo llp': 'Accounting, audit, and advisory services firm',
    'tog uk properties limited': 'Office space and coworking facilities provider',
    'cloudcrossing bvba': 'Cloud infrastructure and IT services provider',
    'amazon web services llc': 'Cloud computing infrastructure and platform services',
    'infosys': 'IT consulting and software development services',
    'linkedin ireland limited': 'Professional networking and recruitment platform',
    'hubspot ireland limited': 'Marketing automation and CRM software platform',
    'google ireland limited': 'Online advertising, cloud services, and workspace tools',
    'workato, inc.': 'Integration and workflow automation platform',
    'microsoft ireland operations limited': 'Enterprise software and cloud computing services',
    'atlassian pty ltd': 'Collaboration and software development tools',
    'figma, inc.': 'Collaborative design and prototyping platform',
    'adobe systems software': 'Creative software and digital marketing tools',
    'docusign': 'Electronic signature and document management platform',
    'smartsheet inc.': 'Work management and collaboration platform',
    'trello': 'Project management and collaboration software',
    'slack technologies limited': 'Team collaboration and messaging platform',
    'zapier inc.': 'Workflow automation and app integration platform',
    # Add more key vendors as needed
}

def get_vendor_description(vendor_name):
    """Get description for a vendor (using lowercase matching)"""
    vendor_lower = vendor_name.lower()

    # Try exact match first
    if vendor_lower in VENDOR_DESCRIPTIONS:
        return VENDOR_DESCRIPTIONS[vendor_lower]

    # Generate generic description based on vendor name
    if 'hotel' in vendor_lower or 'resort' in vendor_lower:
        return 'Hotel accommodation and hospitality services'
    elif 'catering' in vendor_lower or 'kitchen' in vendor_lower:
        return 'Catering and food services provider'
    elif 'restaurant' in vendor_lower or 'cafe' in vendor_lower or 'bar' in vendor_lower:
        return 'Restaurant and dining services'
    elif 'law' in vendor_lower or 'legal' in vendor_lower or 'solicitor' in vendor_lower:
        return 'Legal services and law firm'
    elif 'recruitment' in vendor_lower or 'staffing' in vendor_lower:
        return 'Recruitment and staffing services'
    elif 'insurance' in vendor_lower:
        return 'Insurance and risk management services'
    elif 'accounting' in vendor_lower or 'accountant' in vendor_lower:
        return 'Accounting and financial services'
    elif 'coworking' in vendor_lower or 'office space' in vendor_lower or 'wework' in vendor_lower:
        return 'Coworking and office space provider'
    elif 'event' in vendor_lower:
        return 'Event planning and management services'
    elif 'parking' in vendor_lower:
        return 'Parking facility management services'
    elif 'gym' in vendor_lower or 'fitness' in vendor_lower or 'sports club' in vendor_lower:
        return 'Fitness and recreation services'
    elif 'telecom' in vendor_lower or 'telekom' in vendor_lower or 'mobile' in vendor_lower:
        return 'Telecommunications services provider'
    elif 'cloud' in vendor_lower:
        return 'Cloud infrastructure and services'
    elif 'consulting' in vendor_lower or 'advisory' in vendor_lower:
        return 'Business consulting and advisory services'
    else:
        return 'Business services provider'

def get_recommendation(vendor_name, description):
    """Generate recommendation based on vendor name AND description"""
    vendor_lower = vendor_name.lower()
    desc_lower = description.lower()

    # TERMINATE — Non-essential, discretionary, or easily replaced services
    terminate_indicators = [
        # Individual contractors
        'individual contractor', 'john smith', 'susan lee', 'george anchor',
        'fabiola thistlewhaite', 'stipe piric', 'ansar madovic',
        # Non-essential retail/gifts
        'pink ribbon shop', 'regency hampers', 'cupcake central', 'the plant man',
        # Entertainment (non-core)
        'djs for u', 'paint & fun', 'paint&wine', 'lajnap comedy', 'escape art',
        # Recreation/dining (discretionary)
        'gym', 'fitness center', 'sports club', 'recreation club', 'wine retail',
        'istra wine', 'vivat fina vina', 'notino s.r.o.', 'freepik company',
        'magic mountain saloon', 'pepe\'s italian', 'friends sports club',
        'chamiers recreation', 'p s recreation', 'the cycle gap',
    ]

    # Check description for terminate keywords
    terminate_desc_keywords = [
        'individual contractor', 'gym membership', 'recreation club',
        'sports club', 'wine retail', 'entertainment booking',
        'creative workshop', 'escape room'
    ]

    for keyword in terminate_indicators:
        if keyword in vendor_lower:
            return 'Terminate'

    for keyword in terminate_desc_keywords:
        if keyword in desc_lower:
            return 'Terminate'

    # CONSOLIDATE — Overlapping services, discretionary travel/events, duplicate tools
    consolidate_desc_keywords = [
        # Travel & hospitality (use corporate travel platform instead)
        'hotel', 'resort', 'accommodation', 'hospitality',
        # Food services (consolidate to fewer providers)
        'catering', 'restaurant', 'dining', 'food services', 'meal services',
        # Events (consolidate event vendors)
        'event planning', 'event management', 'conference',
        # Parking (consolidate to fewer providers)
        'parking', 'garage management',
        # Office space (consolidate to primary provider)
        'coworking', 'office space', 'workspace', 'flexible office',
        # Overlapping SaaS/cloud tools
        'saas', 'platform', 'software as a service',
        # Professional services with multiple vendors
        'consulting', 'advisory services', 'recruitment', 'staffing',
        'legal services', 'law firm', 'accounting services',
        # Insurance (consolidate policies)
        'insurance', 'risk management',
        # Telecom (consolidate providers)
        'telecommunications', 'mobile services', 'internet services',
        # Marketing tools (many overlapping)
        'marketing automation', 'sales intelligence', 'digital marketing',
    ]

    for keyword in consolidate_desc_keywords:
        if keyword in desc_lower:
            return 'Consolidate'

    # OPTIMIZE — Core strategic platforms and essential services
    optimize_indicators = [
        # Major strategic platforms (keep but negotiate)
        'salesforce', 'aws', 'amazon web services', 'microsoft', 'google ireland',
        'hubspot', 'linkedin', 'workato', 'atlassian', 'figma', 'adobe',
        'docusign', 'smartsheet', 'trello', 'slack', 'zapier',
        # Critical professional services (Big 4, etc.)
        'bdo llp', 'grant thornton', 'pricewaterhousecoopers', 'rsm uk corporate',
        'houlihan lokey', 'crowe horwath',
        # Core IT/Engineering
        'infosys', 'jetbrains', 'ariba', 'kimble', 'pluralsight',
        # Essential infrastructure
        'dhl', 'fedex', 'british telecommunications',
        # Primary providers
        'navan (tripactions inc)', 'navan, inc', 'cbre limited', 'jones lang lasalle',
        'benefit systems', 'mercer limited', 'pluxee india', 'sodexo', 'granttree limited',
        'lastpass', 'solarwinds', 'uptime robot', 'papertrail',
    ]

    optimize_desc_keywords = [
        'cloud computing infrastructure', 'crm and sales automation',
        'enterprise software', 'workflow automation', 'collaboration and software development',
        'it consulting', 'audit, tax, and consulting', 'investment banking',
        'logistics and international shipping', 'r&d tax credits',
        'password management', 'it management and monitoring',
    ]

    for keyword in optimize_indicators:
        if keyword in vendor_lower:
            return 'Optimize'

    for keyword in optimize_desc_keywords:
        if keyword in desc_lower:
            return 'Optimize'

    # Default logic based on vendor type
    if any(word in vendor_lower for word in ['d.o.o.', 'j.d.o.o.', 'obrt']):
        # Croatian local businesses
        if any(word in desc_lower for word in ['restaurant', 'bar', 'cafe', 'catering', 'food']):
            return 'Consolidate'
        elif any(word in desc_lower for word in ['retail', 'grocery', 'shop']):
            return 'Consolidate'
        else:
            return 'Optimize'

    # Default: Consolidate for most remaining vendors
    return 'Consolidate'

# Print table
print("| Vendor Name | Recommendation |")
print("|-------------|----------------|")

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        vendor_name = row[0]
        description = get_vendor_description(vendor_name)
        recommendation = get_recommendation(vendor_name, description)
        print(f"| {vendor_name} | {recommendation} |")
