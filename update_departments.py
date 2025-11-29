#!/usr/bin/env python3
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
ws = wb.active

# Define department mappings based on user-provided rules
engineering_vendors = {
    'cloudcrossing bvba', 'amazon web services llc', 'infosys', 'workato, inc.',
    'new star networks(nsn)', 'jetbrains s.r.o.', 'smartsheet inc.', 'figma, inc.',
    'pluralsight, llc', 'ariba inc', 'zapier inc.', 'atlassian pty ltd', 'trello',
    'docusign', 'tm forum', 'yoxel, inc', 'radius group, inc',
    'uptime robot service provider ltd', 'epignosis llc'
}

marketing_vendors = {
    'salesforce uk ltd-uk', 'linkedin ireland limited', 'hubspot ireland limited',
    'cognism limited', 'uberflip', 'google ireland limited', 'semrush inc',
    'outreach corporation', 'terrapinn holdings ltd', 'cision pr newswire',
    'lusha', 'mightyhive ltd'
}

finance_vendors = {
    'bdo llp', 'planful, inc.', 'houlihan lokey advisors, llc', 'sage uk limited',
    'pricewaterhousecoopers llp', 'crowe horwath revizija d.o.o.', 'cigna sg',
    'allianz australia workers\' compensation (victoria) limited',
    'mcburneys charted accountants', 'icici lombard gic ltd',
    'australian taxation office (ato)'
}

legal_vendors = {
    'bisley law ltd', 'pinsent masons mpillay llp', 'oâ€™donnell salzano lawyers',
    'the virtual legal counsel ltd', 'klg - kalra legal group', 'induslaw',
    'quadrant law llc', 'curzon green solicitors', 'g s notary public limited',
    'landu law solicitors'
}

def get_department(vendor_name):
    """Determine department based on vendor name"""
    vendor_lower = vendor_name.lower()

    if vendor_lower in engineering_vendors:
        return 'Engineering'
    elif vendor_lower in marketing_vendors:
        return 'Marketing'
    elif vendor_lower in finance_vendors:
        return 'Finance'
    elif vendor_lower in legal_vendors:
        return 'Legal'
    else:
        return 'G&A'

# Update the Department column (column B, index 1)
for row in ws.iter_rows(min_row=2):  # Skip header
    vendor_name = row[0].value
    if vendor_name:
        department = get_department(vendor_name)
        row[1].value = department

# Save the updated workbook
wb.save('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
print("Spreadsheet updated successfully with department classifications.")
