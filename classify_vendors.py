#!/usr/bin/env python3
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('/home/user/Vendor-Analysis-Assessment/Vendor Analysis Assessment - Deeba.xlsx')
ws = wb.active

# Get vendor names from the spreadsheet
vendors = []
for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
    if row[0]:  # If vendor name exists
        vendors.append(row[0])

# Classification rules based on vendor name and business type
def classify_vendor(vendor_name):
    vendor_lower = vendor_name.lower()

    # Legal - must come before checking LLP
    if any(keyword in vendor_lower for keyword in ['law', 'legal', 'solicitor', 'odvjetnicko', 'notary',
                                                     'pinsent masons', 'kilgannon & partners']):
        return 'Legal'

    # Finance - check before generic LLP
    if any(keyword in vendor_lower for keyword in [
        'insurance', 'osiguranje', 'bdo', 'rsm', 'grant thornton', 'pricewaterhouse', 'pwc',
        'chartered accountant', 'finance', 'houlihan lokey', 'vector capital',
        'sage', 'planful', 'collards', 'mcburney', 'shastri', 'mercer limited',
        'crowe horwath', 'tax', 'cigna', 'bupa', 'aetna', 'icare', 'allianz', 'icici lombard',
        'taxation office', 'australian taxation office'
    ]):
        return 'Finance'

    # Marketing - check before Engineering for tools that could be both
    if any(keyword in vendor_lower for keyword in [
        'salesforce', 'linkedin', 'hubspot', 'cognism', 'uberflip', 'google ireland', 'mightyhive',
        'semrush', 'lusha', 'outreach corporation', 'cision', 'terrapinn'
    ]):
        return 'Marketing'

    # Engineering (Cloud, IT, Software Development)
    if any(keyword in vendor_lower for keyword in [
        'aws', 'amazon web services', 'cloud', 'intralinks', 'infosys', 'workato',
        'kimble', 'jetbrains', 'adobe', 'microsoft', 'npm', 'github', 'gitlab',
        'tech solutions', 'it solutions', 'smartsheet', 'trello', 'jira', 'aha!',
        'docusign', 'fastspring', 'ariba', 'tmforum', 'tm forum', 'new star networks',
        'shree info', 'telefonica', 'kryterion', 'yoxel', 'radius group',
        'trending technology', 'epignosis', 'papertrail', 'atlassian', 'zapier',
        'solarwinds', 'figma', 'lastpass', 'pluralsight', 'uptime robot'
    ]):
        return 'Engineering'

    # Support (Customer support, Help desk)
    if any(keyword in vendor_lower for keyword in [
        'peakon', 'zendesk', 'freshdesk', 'intercom', 'support'
    ]):
        return 'Support'

    # G&A (General & Administrative - HR, Travel, Office, Facilities, Recruiting, etc.)
    if any(keyword in vendor_lower for keyword in [
        'navan', 'tripaction', 'properties', 'tower', 'spaces', 'wework', 'office',
        'tog uk', 'zagrebtower', 'innovent spaces', 'weking', 'gpt space', 'recruitment',
        'hr solution', 'accutrainee', 'mason frank', 'cedar recruitment', 'technet',
        'hotel', 'resort', 'catering', 'restaurant', 'travel', 'sodexo', 'benefit systems',
        'studentski', 'recreation', 'gym', 'parking', 'telekom', 'telecom', 'starhub',
        't-mobile', 'vodafone', 'british telecommunications', 'goto', 'slack',
        'konzum', 'ikea', 'transport', 'fakultet', 'university', 'grad ', 'city ',
        'uber', 'office move', 'blink events', 'acclime', 'intertrust', 'cbre',
        'jones lang', 'plus your business', 'work easy', 'backoffice', 'integrated personnel',
        'visalogic', 'green commute', 'pluxee', 'event', 'golubica parking', 'aquila remete',
        'dsv solutions', 'computershare', 'winmaxi tours', 'lunch nutrition', 'food', 'cafe',
        'stipe piric', 'ansar madovic', 'susan lee', 'john smith', 'fabiola', 'george anchor',
        'anchor recruitment'
    ]):
        return 'G&A'

    # Default to G&A for facilities, catering, and general services
    return 'G&A'

# Classify all vendors
classified_vendors = []
for vendor in vendors:
    department = classify_vendor(vendor)
    classified_vendors.append((vendor, department))

# Print table
print("| Vendor Name | Department |")
print("|-------------|------------|")
for vendor, dept in classified_vendors:
    print(f"| {vendor} | {dept} |")
